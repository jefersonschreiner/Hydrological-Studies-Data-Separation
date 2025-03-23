from openpyxl import load_workbook

# Carrega o arquivo Excel
arquivo = load_workbook("DadosEst.xlsx")

# Função para criar abas
def criar_aba_ano(ano, arquivo):
    if str(ano) not in arquivo.sheetnames:  # Converte o ano para string
        nova_aba = arquivo.create_sheet(str(ano))
        return nova_aba, True  # Retorna a nova aba e um indicador de que foi criada
    return arquivo[str(ano)], False  # Retorna a aba existente e um indicador de que já existia

# Função para transferir dados
def transferir(aba_origem, aba_destino, linha_origem, ultima_coluna):
    linha_destino = aba_destino.max_row + 1
    for coluna in range(1, ultima_coluna + 1):
        celula_origem = aba_origem.cell(row=linha_origem, column=coluna)
        celula_destino = aba_destino.cell(row=linha_destino, column=coluna)
        celula_destino.value = celula_origem.value

# Verificações
aba_dados = arquivo["Dados"]
ultima_linha = aba_dados.max_row
ultima_coluna = aba_dados.max_column
print("Última linha:", ultima_linha)
print("Última coluna:", ultima_coluna)

# Copiar cabeçalhos
cabecalhos = []
for coluna in range(1, ultima_coluna + 1):
    cabecalhos.append(aba_dados.cell(row=1, column=coluna).value)

# Funcionamento do Código
for linha in range(2, ultima_linha + 1):
    valor_celula = aba_dados.cell(row=linha, column=1).value

    # Extrair o ano
    if isinstance(valor_celula, int):  # Se for um número inteiro (ano)
        anos = valor_celula
    else:
        if hasattr(valor_celula, 'year'):  # Se for uma data
            anos = valor_celula.year
        else:
            print(f"Valor inválido na linha {linha}: {valor_celula}")
            continue  # Pula para a próxima linha

    # Criar ou acessar a aba do ano
    aba_destino, criada_agora = criar_aba_ano(anos, arquivo)

    # Se a aba acabou de ser criada, copiar os cabeçalhos
    if criada_agora:
        for coluna, cabecalho in enumerate(cabecalhos, start=1):
            aba_destino.cell(row=1, column=coluna).value = cabecalho

    # Transferir os dados para a aba
    transferir(aba_dados, aba_destino, linha, ultima_coluna)

# Salvar o arquivo modificado
arquivo.save("DadosProcessado.xlsx")
print("Se pah deu certo!")